#!/usr/bin/env python3
"""
Generate TAGGD Dashboard Excel Template v5.10.5
Complete structure with all required sheets and columns
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_excel_template():
    """Create the complete Excel template for TAGGD Dashboard"""
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ========================
    # SHEET 1: FY 24-25 Summary
    # ========================
    ws1 = wb.create_sheet("FY 24-25 Summary")
    
    # Headers for FY 24-25 Summary
    summary_headers = [
        "S.No", "Project", "Category", "Measure", "Regional Head", "Region", 
        "Practice Head", "April", "May", "June", "July", "Aug", "Sep", 
        "Oct", "Nov", "Dec", "Jan", "Industry Type "
    ]
    
    ws1.append(summary_headers)
    
    # Format header row
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data for FY 24-25 Summary
    sample_data_fy24 = [
        [1, "Project Alpha", "Contractual SLA", "Response Time", "John Doe", "Mumbai", "Practice A", "Met", "Met", "Not Met", "Met", "Not Reported", "Met", "Met", "Met", "Not Met", "Met", "IT Services"],
        [2, "Project Alpha", "Internal KPI", "Quality Score", "John Doe", "Mumbai", "Practice A", "Met", "Not Met", "Met", "Met", "Met", "NA", "Met", "Not Reported", "Met", "Met", "IT Services"],
        [3, "Project Beta", "Penalty SLA", "Availability", "Jane Smith", "Bangalore", "Practice B", "Met", "Met", "Met", "Not Met", "Met", "Met", "Not Met", "Met", "Met", "NA", "Financial Services"],
        [4, "Project Beta", "Contractual SLA", "Uptime", "Jane Smith", "Bangalore", "Practice B", "Not Met", "Met", "Met", "Met", "Not Reported", "Met", "Met", "Met", "Met", "Met", "Financial Services"],
        [5, "Project Gamma", "Internal KPI", "Customer Satisfaction", "Robert Lee", "Delhi", "Practice C", "Met", "Met", "Met", "Met", "Met", "Met", "Met", "Not Met", "Not Reported", "Met", "Healthcare"],
    ]
    
    for row in sample_data_fy24:
        ws1.append(row)
    
    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column].width = adjusted_width
    
    # ========================
    # SHEET 2: FY 25-26 Summary
    # ========================
    ws2 = wb.create_sheet("FY 25-26 Summary")
    
    ws2.append(summary_headers)
    
    # Format header
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data for FY 25-26 Summary
    sample_data_fy25 = [
        [1, "Project Alpha", "Contractual SLA", "Response Time", "John Doe", "Mumbai", "Practice A", "Met", "Met", "Met", "Not Met", "Met", "Met", "Not Reported", "Met", "Met", "Met", "IT Services"],
        [2, "Project Alpha", "Internal KPI", "Quality Score", "John Doe", "Mumbai", "Practice A", "Met", "Not Met", "Met", "Met", "Not Reported", "Met", "Met", "Met", "Not Met", "NA", "IT Services"],
        [3, "Project Beta", "Penalty SLA", "Availability", "Jane Smith", "Bangalore", "Practice B", "Met", "Met", "Not Met", "Met", "Met", "Met", "Met", "Not Reported", "Met", "Met", "Financial Services"],
        [4, "Project Beta", "Contractual SLA", "Uptime", "Jane Smith", "Bangalore", "Practice B", "Met", "Met", "Met", "Met", "Met", "Not Met", "Met", "Met", "Not Reported", "Met", "Financial Services"],
        [5, "Project Gamma", "Internal KPI", "Customer Satisfaction", "Robert Lee", "Delhi", "Practice C", "Not Met", "Met", "Met", "Met", "Met", "Met", "Met", "Met", "Met", "Not Reported", "Healthcare"],
    ]
    
    for row in sample_data_fy25:
        ws2.append(row)
    
    # Auto-adjust column widths
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws2.column_dimensions[column].width = adjusted_width
    
    # ========================
    # SHEET 3: FY24-25 Not Reported
    # ========================
    ws3 = wb.create_sheet("FY24-25 Not Reported")
    
    # Headers for Not Reported sheets
    not_reported_headers = [
        "S.No", "Project", "Category", "Measure", "Regional Head", "Region", "Practice Head",
        "Apr MET/NOT_MET", "Apr MET/NOT_MET_NotReported",
        "May MET/NOT_MET", "May MET/NOT_MET_NotReported",
        "Jun MET/NOT_MET", "Jun MET/NOT_MET_NotReported",
        "Jul MET/NOT_MET", "Jul MET/NOT_MET_NotReported",
        "Aug MET/NOT_MET", "Aug MET/NOT_MET_NotReported",
        "Sep MET/NOT_MET", "Sep MET/NOT_MET_NotReported",
        "Oct MET/NOT_MET", "Oct MET/NOT_MET_NotReported",
        "Nov MET/NOT_MET", "Nov MET/NOT_MET_NotReported",
        "Dec MET/NOT_MET", "Dec MET/NOT_MET_NotReported",
        "Jan MET/NOT_MET", "Jan MET/NOT_MET_NotReported"
    ]
    
    ws3.append(not_reported_headers)
    
    # Format header
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data for FY24-25 Not Reported
    sample_not_reported_fy24 = [
        [1, "Project Alpha", "Contractual SLA", "Response Time", "John Doe", "Mumbai", "Practice A", 
         "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Met", 0],
        [2, "Project Alpha", "Internal KPI", "Quality Score", "John Doe", "Mumbai", "Practice A",
         "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Met", 0, "NA", 1, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0],
        [3, "Project Beta", "Penalty SLA", "Availability", "Jane Smith", "Bangalore", "Practice B",
         "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "NA", 1],
        [4, "Project Beta", "Contractual SLA", "Uptime", "Jane Smith", "Bangalore", "Practice B",
         "Not Met", 0, "Met", 0, "Met", 0, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0],
        [5, "Project Gamma", "Internal KPI", "Customer Satisfaction", "Robert Lee", "Delhi", "Practice C",
         "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Not Reported", 1, "Met", 0],
    ]
    
    for row in sample_not_reported_fy24:
        ws3.append(row)
    
    # Auto-adjust column widths
    for i, col in enumerate(ws3.columns, 1):
        if i <= 7:  # First 7 columns
            ws3.column_dimensions[col[0].column_letter].width = 20
        else:  # Month columns
            ws3.column_dimensions[col[0].column_letter].width = 15
    
    # ========================
    # SHEET 4: FY25-26 Not Reported
    # ========================
    ws4 = wb.create_sheet("FY25-26 Not Reported")
    
    ws4.append(not_reported_headers)
    
    # Format header
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data for FY25-26 Not Reported
    sample_not_reported_fy25 = [
        [1, "Project Alpha", "Contractual SLA", "Response Time", "John Doe", "Mumbai", "Practice A",
         "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0, "Met", 0],
        [2, "Project Alpha", "Internal KPI", "Quality Score", "John Doe", "Mumbai", "Practice A",
         "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "NA", 1],
        [3, "Project Beta", "Penalty SLA", "Availability", "Jane Smith", "Bangalore", "Practice B",
         "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Not Reported", 1, "Met", 0, "Met", 0],
        [4, "Project Beta", "Contractual SLA", "Uptime", "Jane Smith", "Bangalore", "Practice B",
         "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Not Met", 0, "Met", 0, "Met", 0, "Not Reported", 1, "Met", 0],
        [5, "Project Gamma", "Internal KPI", "Customer Satisfaction", "Robert Lee", "Delhi", "Practice C",
         "Not Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Met", 0, "Not Reported", 1],
    ]
    
    for row in sample_not_reported_fy25:
        ws4.append(row)
    
    # Auto-adjust column widths
    for i, col in enumerate(ws4.columns, 1):
        if i <= 7:  # First 7 columns
            ws4.column_dimensions[col[0].column_letter].width = 20
        else:  # Month columns
            ws4.column_dimensions[col[0].column_letter].width = 15
    
    # ========================
    # SHEET 5: FY 25-26 Metrics Details
    # ========================
    ws5 = wb.create_sheet("FY 25-26 Metrics Details ")  # Note: Space at the end is intentional
    
    # Headers for Metrics Details
    metrics_headers = [
        "S.No", "Project", "Category", "Performance Measure ", "Regional Head", "Region", 
        "Practice Head",
        "Apr MET/NOT_MET", "May MET/NOT_MET", "Jun MET/NOT_MET", "Jul MET/NOT_MET",
        "Aug MET/NOT_MET", "Sep MET/NOT_MET", "Oct MET/NOT_MET", "Nov MET/NOT_MET",
        "Dec MET/NOT_MET", "Jan MET/NOT_MET"
    ]
    
    ws5.append(metrics_headers)
    
    # Format header
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data for Metrics Details
    sample_metrics = [
        [1, "Project Alpha", "Contractual SLA", "Response Time < 5 seconds", "John Doe", "Mumbai", "Practice A", 
         "Met", "Met", "Met", "Not Met", "Met", "Met", "Not Reported", "Met", "Met", "Met"],
        [2, "Project Alpha", "Internal KPI", "Quality Score > 90%", "John Doe", "Mumbai", "Practice A",
         "Met", "Not Met", "Met", "Met", "Not Reported", "Met", "Met", "Met", "Not Met", "NA"],
        [3, "Project Beta", "Penalty SLA", "Availability 99.9%", "Jane Smith", "Bangalore", "Practice B",
         "Met", "Met", "Not Met", "Met", "Met", "Met", "Met", "Not Reported", "Met", "Met"],
        [4, "Project Beta", "Contractual SLA", "Uptime > 99.5%", "Jane Smith", "Bangalore", "Practice B",
         "Met", "Met", "Met", "Met", "Met", "Not Met", "Met", "Met", "Not Reported", "Met"],
        [5, "Project Gamma", "Internal KPI", "Customer Satisfaction > 4.5", "Robert Lee", "Delhi", "Practice C",
         "Not Met", "Met", "Met", "Met", "Met", "Met", "Met", "Met", "Met", "Not Reported"],
    ]
    
    for row in sample_metrics:
        ws5.append(row)
    
    # Auto-adjust column widths
    for col in ws5.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws5.column_dimensions[column].width = adjusted_width
    
    # ========================
    # SHEET 6: Instructions
    # ========================
    ws6 = wb.create_sheet("Instructions")
    
    instructions = [
        ["TAGGD Customer SLA Dashboard - Excel Template v5.10.5"],
        [""],
        ["CRITICAL: Do NOT modify sheet names or column headers!"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["REQUIRED SHEETS (All 5 sheets are mandatory):"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["1. FY 24-25 Summary"],
        ["   - Historical data for fiscal year 2024-2025"],
        ["   - Columns: S.No, Project, Category, Measure, Regional Head, Region, Practice Head,"],
        ["              April, May, June, July, Aug, Sep, Oct, Nov, Dec, Jan, Industry Type "],
        ["   - Values: Met, Not Met, Not Reported, NA"],
        [""],
        ["2. FY 25-26 Summary"],
        ["   - Current data for fiscal year 2025-2026"],
        ["   - Same column structure as FY 24-25 Summary"],
        ["   - Values: Met, Not Met, Not Reported, NA"],
        [""],
        ["3. FY24-25 Not Reported"],
        ["   - Detailed Not Reported tracking for FY 24-25"],
        ["   - Columns: S.No, Project, Category, Measure, Regional Head, Region, Practice Head,"],
        ["              Apr MET/NOT_MET, Apr MET/NOT_MET_NotReported,"],
        ["              May MET/NOT_MET, May MET/NOT_MET_NotReported,"],
        ["              (repeat for Jun, Jul, Aug, Sep, Oct, Nov, Dec, Jan)"],
        ["   - MET/NOT_MET values: Met, Not Met, Not Reported, NA"],
        ["   - NotReported values: 0 (reported) or 1 (not reported)"],
        [""],
        ["4. FY25-26 Not Reported"],
        ["   - Detailed Not Reported tracking for FY 25-26"],
        ["   - Same column structure as FY24-25 Not Reported"],
        ["   - MET/NOT_MET values: Met, Not Met, Not Reported, NA"],
        ["   - NotReported values: 0 (reported) or 1 (not reported)"],
        [""],
        ["5. FY 25-26 Metrics Details "],
        ["   - Detailed performance metrics for drill-down modal"],
        ["   - IMPORTANT: Sheet name ends with a SPACE character"],
        ["   - Columns: S.No, Project, Category, Performance Measure , Regional Head, Region,"],
        ["              Practice Head, Apr MET/NOT_MET, May MET/NOT_MET, ... Jan MET/NOT_MET"],
        ["   - Values: Met, Not Met, Not Reported, NA"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["VALID VALUES FOR STATUS COLUMNS:"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["• Met          - SLA/KPI achieved"],
        ["• Not Met      - SLA/KPI missed"],
        ["• Not Reported - Data not available (excluded from % calculation)"],
        ["• NA           - Not applicable (excluded from % calculation)"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["CATEGORIES (Case-sensitive):"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["• Contractual SLA  - SLAs with contractual obligations"],
        ["• Internal KPI     - Internal performance indicators"],
        ["• Penalty SLA      - SLAs with penalty clauses (subset of Contractual)"],
        [""],
        ["NOTE: Non-Penalty SLA is auto-calculated as Contractual SLA - Penalty SLA"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["MONTH ABBREVIATIONS (Must match exactly):"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec, Jan"],
        [""],
        ["IMPORTANT: Use short month names (Apr, not April)"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["MET % CALCULATION LOGIC:"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["Met % = (Met / (Met + Not Met)) × 100"],
        [""],
        ["CRITICAL: 'Not Reported' and 'NA' are EXCLUDED from the denominator"],
        [""],
        ["Example:"],
        ["  Met: 8, Not Met: 2, Not Reported: 3"],
        ["  Met % = (8 / (8 + 2)) × 100 = 80.0%"],
        ["  (The 3 'Not Reported' entries are not counted)"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["NOT REPORTED TRACKING:"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["In 'Not Reported' sheets:"],
        ["  - Each month has 2 columns: [Month] MET/NOT_MET and [Month] MET/NOT_MET_NotReported"],
        ["  - MET/NOT_MET column: Status value (Met, Not Met, Not Reported, NA)"],
        ["  - NotReported column: Numeric value (0 = reported, 1 = not reported)"],
        [""],
        ["Example row:"],
        ["  Apr MET/NOT_MET: 'Not Reported'"],
        ["  Apr MET/NOT_MET_NotReported: 1"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["EXPECTED TOTALS (For verification):"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["Not Reported Analysis - No Filters:"],
        ["  FY 24-25 Total: 567"],
        ["  FY 25-26 Total: 1,055"],
        [""],
        ["Not Reported Analysis - Month = Oct:"],
        ["  FY 24-25: ~47"],
        ["  FY 25-26: 102"],
        [""],
        ["Not Reported Analysis - Month = Dec:"],
        ["  FY 24-25: ~43"],
        ["  FY 25-26: 149"],
        [""],
        ["Not Reported Analysis - Month = Jan:"],
        ["  FY 24-25: ~45"],
        ["  FY 25-26: 113"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["TIPS:"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["1. Keep sheet names EXACTLY as shown (including spaces)"],
        ["2. Keep column headers EXACTLY as shown (case-sensitive)"],
        ["3. Use only the valid status values listed above"],
        ["4. Use short month names (Apr, not April)"],
        ["5. Ensure 'Industry Type ' column has a trailing space"],
        ["6. Ensure 'Performance Measure ' column has a trailing space"],
        ["7. 'FY 25-26 Metrics Details ' sheet name has a trailing space"],
        ["8. Fill all required columns (no empty cells for key fields)"],
        ["9. S.No should be sequential numbers starting from 1"],
        ["10. Consistent naming: Same project name across all sheets"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
        ["VERSION INFORMATION:"],
        ["═══════════════════════════════════════════════════════════════"],
        [""],
        ["Template Version: v5.10.5"],
        ["Dashboard Version: v5.10.5"],
        ["Created: " + datetime.now().strftime("%Y-%m-%d")],
        ["GitHub: https://github.com/Businessexcellence/SLA-DASHBOARD"],
        ["Commit: 78bf81a"],
        [""],
        ["Password: GoGetter@2026"],
        [""],
        ["═══════════════════════════════════════════════════════════════"],
    ]
    
    for row in instructions:
        ws6.append(row)
    
    # Format instructions sheet
    ws6.column_dimensions['A'].width = 80
    
    # Bold headers
    for row in [1, 3, 6, 38, 52, 60, 67, 75, 88, 101, 115]:
        ws6[f'A{row}'].font = Font(bold=True, size=12)
    
    # Orange highlight for critical notes
    critical_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    for row in [3]:
        ws6[f'A{row}'].fill = critical_fill
        ws6[f'A{row}'].font = Font(bold=True, size=11, color="CC0000")
    
    # Save the workbook
    filename = f"TAGGD_Dashboard_Template_v5.10.5_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created: {filename}")
    print(f"\n📊 Template includes:")
    print(f"   1. FY 24-25 Summary (5 sample rows)")
    print(f"   2. FY 25-26 Summary (5 sample rows)")
    print(f"   3. FY24-25 Not Reported (5 sample rows)")
    print(f"   4. FY25-26 Not Reported (5 sample rows)")
    print(f"   5. FY 25-26 Metrics Details  (5 sample rows)")
    print(f"   6. Instructions (Complete guide)")
    print(f"\n🔑 Dashboard Password: GoGetter@2026")
    print(f"\n📝 Next steps:")
    print(f"   1. Open the Excel file")
    print(f"   2. Read the Instructions sheet")
    print(f"   3. Replace sample data with your real data")
    print(f"   4. Keep column headers EXACTLY as shown")
    print(f"   5. Upload to dashboard at: https://3000-in27j4kvifkpo1odihjj8-b237eb32.sandbox.novita.ai")
    
    return filename

if __name__ == "__main__":
    create_excel_template()
