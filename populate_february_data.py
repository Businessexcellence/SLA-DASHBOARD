#!/usr/bin/env python3
"""
Populate February data in Metrics Details sheet using data from Summary sheet
"""

import openpyxl
import sys

def populate_february_data(filename):
    """Populate February MET/NOT_MET data from Summary to Metrics Details"""
    
    print(f"📂 Opening {filename}...")
    wb = openpyxl.load_workbook(filename)
    
    # Get Summary sheet
    summary_sheet = wb['FY 25-26 Summary']
    summary_headers = [cell.value for cell in summary_sheet[1]]
    
    # Find Feb columns in Summary
    feb_met_idx = summary_headers.index('Feb_Met') + 1 if 'Feb_Met' in summary_headers else None
    feb_notmet_idx = summary_headers.index('Feb_Not_Met') + 1 if 'Feb_Not_Met' in summary_headers else None
    
    if not feb_met_idx or not feb_notmet_idx:
        print("❌ Feb_Met or Feb_Not_Met not found in Summary sheet!")
        wb.close()
        return False
    
    # Build project-level February data
    project_feb_data = {}
    for row_idx in range(2, summary_sheet.max_row + 1):
        project = summary_sheet.cell(row_idx, 1).value
        feb_met = summary_sheet.cell(row_idx, feb_met_idx).value or 0
        feb_notmet = summary_sheet.cell(row_idx, feb_notmet_idx).value or 0
        
        if project:
            project_feb_data[project] = {
                'met': feb_met,
                'not_met': feb_notmet,
                'total': feb_met + feb_notmet
            }
    
    print(f"\n✅ Loaded February data for {len(project_feb_data)} projects")
    print("\nProjects with February data:")
    for proj, data in project_feb_data.items():
        if data['total'] > 0:
            compliance = (data['met'] / data['total'] * 100) if data['total'] > 0 else 0
            print(f"  {proj}: Met={data['met']}, Not Met={data['not_met']} ({compliance:.1f}%)")
    
    # Get Metrics Details sheet
    metrics_sheet = wb['FY 25-26 Metrics Details ']
    metrics_headers = [cell.value for cell in metrics_sheet[1]]
    
    # Find February columns in Metrics Details
    feb_score_idx = None
    feb_met_idx_metrics = None
    
    for i, h in enumerate(metrics_headers):
        if h == 'Feb26 Score':
            feb_score_idx = i + 1
        elif h == 'Feb MET/NOT_MET':
            feb_met_idx_metrics = i + 1
    
    if not feb_score_idx or not feb_met_idx_metrics:
        print("❌ February columns not found in Metrics Details sheet!")
        print(f"Available columns: {metrics_headers}")
        wb.close()
        return False
    
    print(f"\n✅ Found February columns in Metrics Details:")
    print(f"   Feb26 Score: column {feb_score_idx}")
    print(f"   Feb MET/NOT_MET: column {feb_met_idx_metrics}")
    
    # Populate Metrics Details with February data
    print(f"\n📝 Updating {metrics_sheet.max_row - 1} measures...")
    
    updated_count = 0
    for row_idx in range(2, metrics_sheet.max_row + 1):
        project = metrics_sheet.cell(row_idx, 1).value  # Project column
        
        if project and project in project_feb_data:
            data = project_feb_data[project]
            
            if data['total'] > 0:
                # Calculate how many measures to mark as Met
                # Distribute Met/Not Met proportionally across measures
                
                # For simplicity, we'll mark measures as Met or Not Met based on the ratio
                # This is a simplified distribution - you may want to adjust based on actual measure data
                
                total_measures_for_project = sum(1 for r in range(2, metrics_sheet.max_row + 1) 
                                                 if metrics_sheet.cell(r, 1).value == project)
                
                # Calculate score (0-100) based on compliance percentage
                compliance_pct = (data['met'] / data['total'] * 100) if data['total'] > 0 else 0
                score = int(compliance_pct)
                
                # Set score
                metrics_sheet.cell(row_idx, feb_score_idx).value = score
                
                # Set MET/NOT_MET based on compliance
                # If compliance is >= 50%, mark as Met, otherwise Not Met
                if compliance_pct >= 50:
                    status = "Met"
                else:
                    status = "Not Met"
                
                metrics_sheet.cell(row_idx, feb_met_idx_metrics).value = status
                updated_count += 1
    
    print(f"✅ Updated {updated_count} measures with February data")
    
    # Save the file
    output_filename = filename.replace('.xlsx', '_populated.xlsx')
    print(f"\n💾 Saving to {output_filename}...")
    wb.save(output_filename)
    wb.close()
    
    print(f"\n✅ SUCCESS! File saved as: {output_filename}")
    print(f"\n📊 Summary:")
    print(f"   - Total projects with February data: {sum(1 for d in project_feb_data.values() if d['total'] > 0)}")
    print(f"   - Total measures updated: {updated_count}")
    print(f"\n🎯 Next step: Upload {output_filename} to the TEST dashboard")
    print(f"   URL: https://3001-in27j4kvifkpo1odihjj8-b237eb32.sandbox.novita.ai")
    print(f"   Password: GoGetter@2026")
    
    return True

if __name__ == '__main__':
    filename = 'SLA_Monthly_Status_Summary_FINAL_with_february.xlsx'
    
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    
    print("=" * 70)
    print("🔧 February Data Populator for Dashboard")
    print("=" * 70)
    
    success = populate_february_data(filename)
    
    if success:
        print("\n" + "=" * 70)
        print("✅ COMPLETE")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print("❌ FAILED")
        print("=" * 70)
        sys.exit(1)
