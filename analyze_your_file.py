#!/usr/bin/env python3
"""
Quick data summary from your Excel file
"""

import openpyxl

print("\n" + "="*70)
print("  QUICK DATA SUMMARY - SLA_Monthly_Status_Summary_FINAL.xlsx")
print("="*70)

wb = openpyxl.load_workbook('SLA_Monthly_Status_Summary_FINAL.xlsx', data_only=True)

# FY 25-26 Summary
ws = wb['FY 25-26 Summary']
print("\n📊 FY 25-26 Summary Sheet")
print("-" * 70)
print(f"   Total rows: {ws.max_row - 1} (excluding header)")

# Get Regional Head column
rh_col = 5  # Based on analysis above

# Count by Regional Head
regional_heads = {}
regions = {}
practices = {}

for row in range(2, ws.max_row + 1):
    rh = ws.cell(row=row, column=5).value  # Regional Head
    region = ws.cell(row=row, column=3).value  # Region
    practice = ws.cell(row=row, column=4).value  # Practice Head
    
    if rh:
        rh_str = str(rh).strip()
        regional_heads[rh_str] = regional_heads.get(rh_str, 0) + 1
    
    if region:
        region_str = str(region).strip()
        regions[region_str] = regions.get(region_str, 0) + 1
    
    if practice:
        practice_str = str(practice).strip()
        practices[practice_str] = practices.get(practice_str, 0) + 1

print("\n   🎯 Regional Head Breakdown:")
for rh, count in sorted(regional_heads.items()):
    percentage = (count / (ws.max_row - 1)) * 100
    print(f"      • {rh}: {count} projects ({percentage:.1f}%)")

print("\n   📍 Top 5 Regions:")
for region, count in sorted(regions.items(), key=lambda x: x[1], reverse=True)[:5]:
    print(f"      • {region}: {count} projects")

print("\n   👥 Top 5 Practice Heads:")
for practice, count in sorted(practices.items(), key=lambda x: x[1], reverse=True)[:5]:
    print(f"      • {practice}: {count} projects")

# Check for "North"
print("\n   🔍 Searching for 'North' in Regional Head column...")
north_found = False
for rh in regional_heads.keys():
    if 'north' in rh.lower():
        print(f"      ✅ FOUND: '{rh}'")
        north_found = True

if not north_found:
    print(f"      ❌ NO 'North' entries found")
    print(f"      💡 Your file contains: {', '.join(list(regional_heads.keys()))}")

# FY 25-26 Metrics Details
ws_metrics = wb['FY 25-26 Metrics Details ']
print("\n📊 FY 25-26 Metrics Details Sheet")
print("-" * 70)
print(f"   Total measures: {ws_metrics.max_row - 1}")

# Count by category
categories = {}
for row in range(2, min(ws_metrics.max_row + 1, 1000)):
    category = ws_metrics.cell(row=row, column=5).value  # Category column
    if category:
        cat_str = str(category).strip()
        categories[cat_str] = categories.get(cat_str, 0) + 1

print("\n   📋 Category Breakdown:")
for cat, count in sorted(categories.items()):
    percentage = (count / (ws_metrics.max_row - 1)) * 100
    print(f"      • {cat}: {count} measures ({percentage:.1f}%)")

# Not Reported Analysis
ws_nr = wb['FY25-26 Not Reported']
print("\n📊 FY25-26 Not Reported Sheet")
print("-" * 70)
print(f"   Total rows: {ws_nr.max_row - 1}")

# Calculate total not reported
total_not_reported = 0
for row in range(2, ws_nr.max_row + 1):
    for col in range(5, ws_nr.max_column + 1):  # NotReported columns start at 5
        val = ws_nr.cell(row=row, column=col).value
        if val:
            try:
                total_not_reported += int(val)
            except:
                pass

print(f"   📉 Total Not Reported count: {total_not_reported}")

print("\n" + "="*70)
print("  🎯 WHAT YOU WILL SEE ON DASHBOARD AFTER UPLOAD")
print("="*70)
print("\n   1. Overview Dashboard:")
print(f"      • Total Projects: ~{ws.max_row - 1}")
print(f"      • Total Measures: ~{ws_metrics.max_row - 1}")

print("\n   2. Filters Available:")
print("      • Regional Head dropdown will show:")
for rh in sorted(regional_heads.keys()):
    print(f"         - {rh}")
print("      • Region dropdown will show: (top 5)")
for region in sorted(regions.keys(), key=lambda x: regions[x], reverse=True)[:5]:
    print(f"         - {region}")

print("\n   3. Not Reported Analysis:")
print(f"      • FY 25-26 Total: ~{total_not_reported}")

print("\n   ⚠️  IMPORTANT NOTE:")
if not north_found:
    print("      Dashboard will NOT show 'North' because it's not in your file!")
    print(f"      Your file only contains: {', '.join(list(regional_heads.keys()))}")
else:
    print("      Dashboard WILL show 'North' entries")

print("\n" + "="*70)
print("  ✅ DATA REFRESH VERIFICATION TEST")
print("="*70)
print("\n   To test if data refresh is working:")
print("\n   1. Upload this file to TEST dashboard")
print("   2. Console should show:")
print(f"      📊 FY 25-26 records: {ws.max_row - 1}")
print(f"      📊 FY25-26 Not Reported records: {ws_nr.max_row - 1}")
print(f"      📊 Metrics Details records: {ws_metrics.max_row - 1}")
print("\n   3. Regional Head filter should display:")
for rh in sorted(regional_heads.keys()):
    print(f"      ☐ {rh}")
print("\n   4. If you see these exact values → DATA REFRESH IS WORKING! ✅")
print("\n" + "="*70 + "\n")
