#!/usr/bin/env python3
"""
Add February columns to FY 25-26 Metrics Details sheet
This will add the missing Feb26 Score and Feb MET/NOT_MET columns
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import sys

def add_february_columns(filename):
    """Add February columns to Metrics Details sheet"""
    
    print(f"📂 Opening {filename}...")
    wb = openpyxl.load_workbook(filename)
    
    # Check if Metrics Details sheet exists
    sheet_name = 'FY 25-26 Metrics Details '  # Note the trailing space
    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        wb.close()
        return False
    
    sheet = wb[sheet_name]
    print(f"✅ Found sheet: {sheet_name}")
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    print(f"\n📋 Current headers ({len(headers)} columns): {headers[:10]}...")
    
    # Check if February columns already exist
    if 'Feb26 Score' in headers or 'Feb MET/NOT_MET' in headers:
        print("⚠️  February columns already exist!")
        wb.close()
        return True
    
    # Find the position after Jan columns
    jan_score_idx = None
    jan_met_idx = None
    
    for i, h in enumerate(headers):
        if h == 'Jan26 Score':
            jan_score_idx = i
        elif h == 'Jan MET/NOT_MET':
            jan_met_idx = i
    
    if jan_score_idx is None:
        print("❌ Could not find 'Jan26 Score' column!")
        print("Looking for last month column instead...")
        
        # Find the last column with data
        last_col_idx = len(headers)
        print(f"Will add February columns after column {last_col_idx}")
    else:
        last_col_idx = max(jan_score_idx, jan_met_idx) if jan_met_idx else jan_score_idx
        print(f"✅ Found January columns at positions {jan_score_idx}, {jan_met_idx}")
        print(f"Will add February columns after column {last_col_idx + 1}")
    
    # Insert two new columns after the last month
    insert_col = last_col_idx + 2  # Excel columns are 1-indexed
    
    print(f"\n➕ Inserting 2 new columns at position {insert_col}...")
    
    # Insert Feb26 Score column
    sheet.insert_cols(insert_col)
    feb_score_cell = sheet.cell(1, insert_col)
    feb_score_cell.value = "Feb26 Score"
    feb_score_cell.font = Font(bold=True, color="FFFFFF")
    feb_score_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    feb_score_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Insert Feb MET/NOT_MET column
    sheet.insert_cols(insert_col + 1)
    feb_met_cell = sheet.cell(1, insert_col + 1)
    feb_met_cell.value = "Feb MET/NOT_MET"
    feb_met_cell.font = Font(bold=True, color="FFFFFF")
    feb_met_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    feb_met_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print("✅ Added column headers:")
    print(f"   Column {insert_col}: Feb26 Score")
    print(f"   Column {insert_col + 1}: Feb MET/NOT_MET")
    
    # Fill in sample data (you'll need to replace this with actual data)
    print("\n📝 Filling sample data...")
    
    for row_idx in range(2, sheet.max_row + 1):
        # For Feb26 Score, default to 0
        sheet.cell(row_idx, insert_col).value = 0
        sheet.cell(row_idx, insert_col).alignment = Alignment(horizontal='center')
        
        # For Feb MET/NOT_MET, default to "Not Reported"
        sheet.cell(row_idx, insert_col + 1).value = "Not Reported"
        sheet.cell(row_idx, insert_col + 1).alignment = Alignment(horizontal='center')
    
    # Save the file
    output_filename = filename.replace('.xlsx', '_with_february.xlsx')
    print(f"\n💾 Saving to {output_filename}...")
    wb.save(output_filename)
    wb.close()
    
    print(f"\n✅ SUCCESS! File saved as: {output_filename}")
    print(f"\n⚠️  IMPORTANT: The script added default values:")
    print(f"   - Feb26 Score: 0 (placeholder)")
    print(f"   - Feb MET/NOT_MET: 'Not Reported' (placeholder)")
    print(f"\n📝 Next steps:")
    print(f"   1. Open {output_filename}")
    print(f"   2. Update the February data with actual Met/Not Met values")
    print(f"   3. Upload to the dashboard")
    
    return True

if __name__ == '__main__':
    filename = 'SLA_Monthly_Status_Summary_FINAL.xlsx'
    
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    
    print("=" * 70)
    print("🔧 February Column Adder for Dashboard")
    print("=" * 70)
    
    success = add_february_columns(filename)
    
    if success:
        print("\n" + "=" * 70)
        print("✅ COMPLETE")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print("❌ FAILED")
        print("=" * 70)
        sys.exit(1)
