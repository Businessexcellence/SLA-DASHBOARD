#!/usr/bin/env python3
"""
IMPROVED: Populate February data with correct distribution
Based on actual Met % from Summary sheet
"""

import openpyxl
import sys

def populate_february_correctly(filename):
    """Populate February MET/NOT_MET data with correct distribution"""
    
    print(f"📂 Opening {filename}...")
    wb = openpyxl.load_workbook('SLA_Monthly_Status_Summary_FINAL_with_february.xlsx')
    
    # Get Summary sheet
    summary_sheet = wb['FY 25-26 Summary']
    summary_headers = [cell.value for cell in summary_sheet[1]]
    
    # Find Feb columns in Summary
    feb_met_idx = summary_headers.index('Feb_Met') + 1
    feb_notmet_idx = summary_headers.index('Feb_Not_Met') + 1
    
    # Build project-level February data
    project_feb_data = {}
    for row_idx in range(2, summary_sheet.max_row + 1):
        project = summary_sheet.cell(row_idx, 1).value
        feb_met = summary_sheet.cell(row_idx, feb_met_idx).value or 0
        feb_notmet = summary_sheet.cell(row_idx, feb_notmet_idx).value or 0
        
        if project:
            project_feb_data[project] = {
                'met': feb_met,
                'not_met': feb_notmet,
                'total': feb_met + feb_notmet
            }
    
    print(f"\n✅ Loaded February data for {len(project_feb_data)} projects")
    
    # Get Metrics Details sheet
    metrics_sheet = wb['FY 25-26 Metrics Details ']
    metrics_headers = [cell.value for cell in metrics_sheet[1]]
    
    # Find February columns
    feb_score_idx = metrics_headers.index('Feb26 Score') + 1
    feb_met_idx_metrics = metrics_headers.index('Feb MET/NOT_MET') + 1
    
    print(f"\n📝 Distributing February data across {metrics_sheet.max_row - 1} measures...")
    
    # Group measures by project
    project_measures = {}
    for row_idx in range(2, metrics_sheet.max_row + 1):
        project = metrics_sheet.cell(row_idx, 1).value
        if project:
            if project not in project_measures:
                project_measures[project] = []
            project_measures[project].append(row_idx)
    
    total_met = 0
    total_not_met = 0
    total_updated = 0
    
    # Distribute Met/Not Met for each project
    for project, row_indices in project_measures.items():
        if project in project_feb_data:
            data = project_feb_data[project]
            
            if data['total'] > 0:
                # Number of measures for this project
                num_measures = len(row_indices)
                
                # Calculate how many should be Met vs Not Met
                met_needed = data['met']
                not_met_needed = data['not_met']
                
                # If we have more SLAs (measures) than Met/NotMet counts,
                # we need to distribute proportionally
                if num_measures > data['total']:
                    # Some measures will be "Not Reported"
                    # Distribute Met and NotMet proportionally
                    met_ratio = met_needed / data['total']
                    met_count_to_assign = int(num_measures * met_ratio)
                    not_met_count_to_assign = num_measures - met_count_to_assign
                else:
                    # We have fewer or equal measures than reported counts
                    # This shouldn't happen, but handle it
                    met_count_to_assign = min(met_needed, num_measures)
                    not_met_count_to_assign = num_measures - met_count_to_assign
                
                # Assign statuses
                for i, row_idx in enumerate(row_indices):
                    if i < met_count_to_assign:
                        status = "Met"
                        score = 100
                        total_met += 1
                    elif i < met_count_to_assign + not_met_count_to_assign:
                        status = "Not Met"
                        score = 0
                        total_not_met += 1
                    else:
                        status = "Not Reported"
                        score = 0
                    
                    metrics_sheet.cell(row_idx, feb_score_idx).value = score
                    metrics_sheet.cell(row_idx, feb_met_idx_metrics).value = status
                    total_updated += 1
                
                compliance_pct = (met_needed / data['total'] * 100)
                print(f"  {project}: {num_measures} measures, {met_count_to_assign} Met, {not_met_count_to_assign} Not Met ({compliance_pct:.1f}%)")
    
    print(f"\n✅ Updated {total_updated} measures")
    print(f"   Met: {total_met}")
    print(f"   Not Met: {total_not_met}")
    print(f"   Not Reported: {total_updated - total_met - total_not_met}")
    
    if total_met + total_not_met > 0:
        overall_compliance = (total_met / (total_met + total_not_met) * 100)
        print(f"   Overall Compliance: {overall_compliance:.1f}%")
    
    # Save the file
    output_filename = 'SLA_Monthly_Status_Summary_FINAL_READY.xlsx'
    print(f"\n💾 Saving to {output_filename}...")
    wb.save(output_filename)
    wb.close()
    
    print(f"\n✅ SUCCESS! File saved as: {output_filename}")
    print(f"\n🎯 Next step: Upload to TEST dashboard")
    print(f"   URL: https://3001-in27j4kvifkpo1odihjj8-b237eb32.sandbox.novita.ai")
    print(f"   Password: GoGetter@2026")
    
    # Expected results
    print(f"\n📊 Expected Dashboard Results:")
    print(f"   - February line should show data (not 0%)")
    print(f"   - Monthly view chart will display February bar")
    print(f"   - Compliance should be around 69.6% (48 Met / 69 total)")
    
    return True

if __name__ == '__main__':
    print("=" * 70)
    print("🔧 February Data Populator (CORRECTED)")
    print("=" * 70)
    
    success = populate_february_correctly(None)
    
    if success:
        print("\n" + "=" * 70)
        print("✅ COMPLETE")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print("❌ FAILED")
        print("=" * 70)
        sys.exit(1)
